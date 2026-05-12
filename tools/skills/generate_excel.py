"""Excel generation tool — agent can call this to create spreadsheet files."""

import json as _json
from llm.base import ToolDef


TOOL_SCHEMA = {
    "name": "generate_excel",
    "description": "Generate an Excel (.xlsx) spreadsheet file from data. Input a JSON array of objects where each object is a row with column-name keys. Use this when the user asks to export data, create a spreadsheet, or generate an Excel report.",
    "parameters": {
        "type": "object",
        "properties": {
            "data": {
                "type": "array",
                "items": {"type": "object"},
                "description": "Array of row objects, e.g. [{\"Keyword\":\"test\",\"Volume\":1000}]",
            },
            "sheet_name": {"type": "string", "description": "Sheet name", "default": "Sheet1"},
        },
        "required": ["data"],
    },
}


def make_tool() -> ToolDef:
    async def handler(data: list[dict], sheet_name: str = "Sheet1") -> str:
        try:
            import openpyxl
            from pathlib import Path
        except ImportError:
            return "Error: openpyxl not installed. Run: pip install openpyxl"

        if not data:
            return "Error: no data provided. Pass a JSON array of row objects."

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Collect all unique headers across all rows
        headers = list(data[0].keys())
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                val = row_data.get(key, "")
                # Convert non-serializable types
                if isinstance(val, (list, dict)):
                    val = _json.dumps(val, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=val)

        out_dir = Path("data/exports")
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / "export.xlsx"
        wb.save(str(out_path))

        return (
            f"Excel file generated: {out_path.name} ({ws.max_row} rows × {ws.max_column} columns). "
            f"Download: /api/skills/files/{out_path.name}"
        )

    return ToolDef(
        name=TOOL_SCHEMA["name"],
        description=TOOL_SCHEMA["description"],
        parameters=TOOL_SCHEMA["parameters"],
        handler=handler,
    )
